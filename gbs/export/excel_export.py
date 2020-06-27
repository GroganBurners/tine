import logging
import string
from decimal import Decimal
from io import BytesIO
from itertools import chain
from operator import attrgetter

from gbs.models import Expense, Invoice
from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.styles import Fill, Font
from openpyxl.styles.borders import Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

logger = logging.getLogger(__name__)
CURR_FORMAT = "€#,##"


def add_header_row(ws, row="2"):
    ws.append([])
    ws.append(
        [
            "",
            "Date",
            "Details",
            "Company",
            "Money Out",
            "Money In",
            "VAT Due In",
            "VAT Due Out",
            "VAT Total",
            "Total",
        ]
    )
    font = Font(name="Calibri", size=11, bold=True)
    for cell_no in ["B", "C", "D", "E", "F", "G", "H", "I", "J"]:
        cell = ws[cell_no + row]
        cell.font = font


def apply_border_format(ws, row, column, style="thin"):
    thin_border = Border(
        left=Side(style=style),
        right=Side(style=style),
        top=Side(style=style),
        bottom=Side(style=style),
    )
    ws.cell(row=row, column=column).border = thin_border


def print_total_row(ws, row_num):
    ws["B" + str(row_num)] = "Total"
    for let in list(string.ascii_uppercase[4:10]):
        start_cell = f"{let}{row_num}"
        ws[start_cell].number_format = CURR_FORMAT
        if let == "I":
            ws[start_cell] = f"=H{row_num}-G{row_num}"
        elif let == "J":
            ws[start_cell] = f"=J{row_num-1}"
        else:
            ws[start_cell] = f"=SUM({let}2:{let}{row_num-1})"


def adjust_column_widths(ws):
    for col in ws.columns:
        max_length = 0
        column = col[0].column  # Get the column name
        for cell in col:
            try:  # Necessary to avoid error on empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except BaseException:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width


def get_invoices_expenses():
    invoices = Invoice.objects.filter(cash=False)
    expenses = Expense.objects.filter(cash=False)
    result_list = sorted(list(chain(invoices, expenses)), key=attrgetter("date"))
    return result_list


def print_invoice_expense(ws, row_num):
    total = Decimal(0)
    for res in get_invoices_expenses():
        items = res.items.all()
        row_num = row_num + 1
        formula = f"=IF(ISNUMBER(J{row_num-1}),J{row_num-1}-F{row_num},F{row_num})+G{row_num}-E{row_num}-H{row_num}"
        if type(res) == Invoice:
            total = total + res.total
            row = [
                "",
                res.date.strftime("%d/%m/%Y"),
                items[0].description,
                res.customer.name,
                "",
                res.total,
                "",
                res.total_vat,
                "",
                formula,
            ]
            ws.append(row)
            ws["F" + str(row_num)].number_format = CURR_FORMAT
            ws["H" + str(row_num)].number_format = CURR_FORMAT
            ws["J" + str(row_num)].number_format = CURR_FORMAT
        else:
            total = total - res.total
            row = [
                "",
                res.date.strftime("%d/%m/%Y"),
                items[0].description,
                res.supplier.name,
                res.total,
                "",
                res.total_vat,
                "",
                "",
                formula,
            ]
            ws.append(row)
            ws["E" + str(row_num)].number_format = CURR_FORMAT
            ws["G" + str(row_num)].number_format = CURR_FORMAT
            ws["J" + str(row_num)].number_format = CURR_FORMAT
    return row_num


def export_finances():
    wb = Workbook()
    ws = wb.active
    ws.title = "FinanceSheet2018"

    # Sheet header, first row
    row_num = 2

    add_header_row(ws)
    row_num = print_invoice_expense(ws, row_num)
    row_num = row_num + 1
    print_total_row(ws, row_num)

    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToHeight = False
    ws.page_setup.fitToHeight = 0
    ws.page_setup.fitToWidth = 1

    # tab = Table(displayName="Table1", ref="A1:I4")
    # style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
    #                         showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    # tab.tableStyleInfo = style
    # ws.add_table(tab)
    apply_border_format(ws, 3, 2)
    adjust_column_widths(ws)

    buffer = BytesIO()
    wb.save(buffer)
    xls = buffer.getvalue()
    buffer.close()
    return xls
