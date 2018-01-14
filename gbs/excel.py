from .models import Invoice, Expense
from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.styles import Font, Fill
from openpyxl.styles.borders import Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from decimal import Decimal
from io import BytesIO
from itertools import chain
from operator import attrgetter
import string
import logging

logger = logging.getLogger(__name__)

def apply_border_format(ws, row, column, style='thin'):
    thin_border = Border(left=Side(style=style), 
                     right=Side(style=style), 
                     top=Side(style=style), 
                     bottom=Side(style=style))
    ws.cell(row=row, column=column).border = thin_border

def print_total_row(ws, row_num):
    ws["A"+str(row_num)] = "Total"
    for let in list(string.ascii_uppercase[3:9]):
        start_cell = f'{let}{row_num}'
        ws[start_cell] = f"=SUM({let}2:{let}{row_num-1})"

def adjust_column_widths(ws):
    for col in ws.columns:
     max_length = 0
     column = col[0].column # Get the column name
     for cell in col:
         try: # Necessary to avoid error on empty cells
             if len(str(cell.value)) > max_length:
                 max_length = len(cell.value)
         except:
             pass
     adjusted_width = (max_length + 2) * 1.2
     ws.column_dimensions[column].width = adjusted_width

def export_finances():
    wb = Workbook()
    ws = wb.active
    ws.title = "FinanceSheet2018"

    # Sheet header, first row
    row_num = 1
    ws.append(['Date', 'Details', 'Company', 'Money Out', 'Money In', 'VAT Due In', 'VAT Due Out', 'VAT Total', 'Total' ])
    adjust_column_widths(ws)

    invoices = Invoice.objects.all()
    expenses = Expense.objects.all()
    logger.error(invoices)
    logger.error(expenses)
    result_list = sorted(list(chain(invoices, expenses)), key=attrgetter('date'))
    logger.error(result_list)

    total = Decimal(0)

    for res in result_list:
        items = res.items.all()
        if type(res) == Invoice:
            total = total + res.total()
            row = [res.date, items[0].description, 'Customer', '', res.total(), '', res.total_vat(), '', total]
            ws.append(row)
            row_num = row_num + 1
        else:
            total = total - res.total()
            row = [res.date, items[0].description, 'SUPPLIER', res.total(), '', res.total_vat(), '', '', total]
            ws.append(row)
            row_num = row_num + 1

    row_num = row_num + 1
    print_total_row(ws, row_num)

    # tab = Table(displayName="Table1", ref="A1:I4")
    # style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
    #                         showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    # tab.tableStyleInfo = style
    # ws.add_table(tab)
    apply_border_format(ws,3,2)

    buffer = BytesIO()
    wb.save(buffer)
    xls = buffer.getvalue()
    buffer.close()
    return xls
